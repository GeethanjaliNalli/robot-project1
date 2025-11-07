from datetime import date
from dateutil.relativedelta import relativedelta
import time
import random
import openpyxl

class CustomLibrary(object):

    def __init__(self):
        pass
    
    @property
    def _driver(self):
        return self._sel_lib.driver
    
    def get_ms_excel_row_values_into_dictionary_based_on_key(self, filepath, keyName, sheetName=None):
        """Reads Excel and returns a row as a dictionary based on the given key."""
        workbook = openpyxl.load_workbook(filepath)
        sheet_names = workbook.sheetnames
        result_dict = {}
        if sheetName is None:
            sheetName = sheet_names[0]
        if not self.validate_the_sheet_in_ms_excel_file(filepath, sheetName):
            return result_dict
        worksheet = workbook[sheetName]
        rows = list(worksheet.iter_rows(values_only=True))
        headers = rows[0]
        for row in rows[1:]:
            if str(row[0]) != str(keyName):
                continue
            for idx, cell_data in enumerate(row):
                if cell_data is None or str(cell_data).strip() == "":
                    continue
                cell_data = self.get_unique_test_data(cell_data)
                result_dict[str(headers[idx])] = str(cell_data)
        return result_dict

    def validate_the_sheet_in_ms_excel_file(self, filepath, sheetName):
        """Validates if a given sheet exists in the Excel file."""
        try:
            workbook = openpyxl.load_workbook(filepath)
            return sheetName.lower() in (s.lower() for s in workbook.sheetnames)
        except FileNotFoundError:
            print(f"Error: File not found: {filepath}")
            return False
        except Exception as e:
            print(f"Error validating sheet: {e}")
            return False

    def get_unique_test_data(self, testdata):
        """Replaces the word 'UNIQUE' with a unique timestamp string."""
        random_number = f'{random.randint(0, 999999):06d}'
        timestamp = time.strftime("%Y%m%d%H%M%S") + random_number
        for keyword in ["UNIQUE", "Unique", "unique"]:
            testdata = str(testdata).replace(keyword, timestamp)
        return testdata
